from datetime import timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from PyQt5.QtCore import Qt, QVariant
from PyQt5.QtGui import QColor
from PyQt5.QtWidgets import QTableWidgetItem

from db.models import Event
from db.utils import (get_error_tests_for_period, get_reports_for_period,
                      get_session, get_tests_for_period, get_unarchived_tests)
from gui.event_codes import get_codes_dictionary, get_error_dictionary

try:
    codes_dictionary = get_codes_dictionary()
    errors_dictionary = get_error_dictionary()
except Exception:
    print('ошибка')


BASE_DIR = Path(__file__).parent.parent


def add_rows_journal_table(table, events):
    """Заполняет таблицу журнала событий."""
    for event in events:
        table.insertRow(0)
        color = QColor(codes_dictionary[event.event_code][1])
        number = QTableWidgetItem()
        number.setData(Qt.EditRole, QVariant(int(event.number)))
        table.setItem(0, 0, number)
        datetime = QTableWidgetItem()
        datetime.setData(Qt.EditRole, QVariant(event.date_time))
        table.setItem(0, 1, QTableWidgetItem(str(event.date_time.strftime('%Y.%m.%d %H:%M:%S'))))
        kl = QTableWidgetItem()
        kl.setData(Qt.EditRole, QVariant(event.kl))
        table.setItem(0, 2, kl)
        entry_number = QTableWidgetItem()
        entry_number.setData(Qt.EditRole, QVariant(event.entry_number))
        table.setItem(0, 3, entry_number)
        au_address = QTableWidgetItem()
        au_address.setData(Qt.EditRole, QVariant(event.au_address))
        table.setItem(0, 4, au_address)
        event_code = QTableWidgetItem()
        event_code.setData(Qt.EditRole, QVariant(event.event_code))
        table.setItem(0, 5, event_code)
        addit_param = QTableWidgetItem()
        addit_param.setData(Qt.EditRole, QVariant(event.addit_param))
        table.setItem(0, 6, addit_param)
        table.setItem(0, 7, QTableWidgetItem(str(event.description)))

        for j in range(8):
            table.item(0, j).setBackground(color)


def add_rows_main_table(table, tests):
    """Заполняет таблицу результатов тестов."""
    for test in tests:
        params = [
            test.kl, test.exit_code_1, test.exit_code_2, test.exit_code_3, test.ready_time,
            test.device_version, test.sb_duration, test.electricity_sb, test.average_electricity,
            test.temperature, test.Uv, test.Us, test.Un, test.Cv, test.Cs, test.Cn,
            test.charge_electricity_current, test.discharge_rate_upper_ionistor,
            test.discharge_rate_middle_ionistor, test.discharge_rate_lower_ionistor,
            test.reserve_1, test.reserve_2, test.description_1
        ]
        table.insertRow(0)
        number = QTableWidgetItem()
        number.setData(Qt.EditRole, QVariant(int(test.number)))
        table.setItem(0, 0, number)
        time = QTableWidgetItem()
        time.setData(Qt.EditRole, QVariant(test.time))
        table.setItem(0, 1, QTableWidgetItem(str(test.time.strftime('%H:%M:%S'))))

        i = 2
        for param in params[:5]:
            item = QTableWidgetItem()
            item.setData(Qt.EditRole, QVariant(param))
            table.setItem(0, i, item)
            i += 1

        # если код 1.2 == 7.3 или 7.4 или 9.2 или 9.3
        # тогда к значению кода 3 нужно дописать нули, если его длина меньше 3 символов
        str_code_1_2 = f'{test.exit_code_1}{test.exit_code_2}'
        if str_code_1_2 in ('73', '74', '92', '93'):
            str_code_3 = '0' * (3 - len(str(test.exit_code_3))) + str(test.exit_code_3)
            code_3_item = QTableWidgetItem()
            code_3_item.setData(Qt.EditRole, QVariant(str_code_3))
            table.setItem(0, 5, code_3_item)

        device_type = QTableWidgetItem()
        device_type.setData(Qt.EditRole, QVariant(test.device_type))
        table.setItem(0, 7, device_type)

        i = 8
        for param in params[5:]:
            item = QTableWidgetItem()
            item.setData(Qt.EditRole, QVariant(param))
            table.setItem(0, i, item)
            i += 1

        description_2 = QTableWidgetItem()
        if type(test.description_2) == int:
            description_2.setData(Qt.EditRole, QVariant(test.description_2))
        elif test.description_2.isdigit():
            description_2.setData(Qt.EditRole, QVariant(int(test.description_2)))
        else:
            description_2.setData(Qt.EditRole, QVariant(test.description_2))
        table.setItem(0, i, description_2)

        if 1 <= test.exit_code_1 <= 19:
            error_color = 'red'
        elif test.exit_code_1 == 20:
            error_color = 'purple'
        elif test.exit_code_1 >= 100:
            error_color = 'blue'
        else:
            error_color = 'white'

        color = QColor(error_color)

        for j in range(27):
            table.item(0, j).setBackground(color)


def write_journal_to_file(filename, session=get_session()):
    """Записывает журнал событий в эксель-файл."""
    new_workbook = Workbook()
    new_workbook.save(filename)
    workbook = load_workbook(filename)
    sheet = workbook.worksheets[0]
    columns_names = ['№', 'Дата и время', 'КЛ', 'Номер входа', 'АУ', 'Код события', 'Доп. параметр', 'Описание']
    sheet.append(columns_names)
    for event in session.query(Event).all():
        line = [
            event.number, event.date_time.strftime('%Y.%m.%d %H:%M:%S'), event.kl,
            event.entry_number, event.au_address, event.event_code,
            event.addit_param, event.description]
        sheet.append(line)
    sheet.column_dimensions["B"].width = 20
    for col in ["D", "F", "G"]:
        sheet.column_dimensions[col].width = 15
    sheet.column_dimensions["H"].width = 80
    for row in list(sheet.rows):
        for cell in row[:7]:
            cell.alignment = Alignment(horizontal='center')
    sheet['H1'].alignment = Alignment(horizontal='center')
    workbook.save(filename)
    workbook.close()
    session.close()


def write_tests_results_to_file(filename, period=None, is_error_flag=False):
    """Записывает результаты тестов в 'эксель-файл."""
    new_workbook = Workbook()
    new_workbook.save(filename)
    workbook = load_workbook(filename)
    sheet = workbook.worksheets[0]
    sheet.merge_cells('A1:A2')
    sheet.merge_cells('B1:B2')
    sheet.merge_cells('C1:C2')
    sheet.merge_cells('D1:D2')
    sheet.merge_cells('E1:E2')
    sheet.merge_cells('F1:F2')
    sheet.merge_cells('H1:L1')
    sheet.merge_cells('N1:P1')
    sheet.merge_cells('A1:A2')
    sheet.merge_cells('Q1:S1')
    sheet.merge_cells('U1:W1')
    sheet.merge_cells('X1:X2')
    sheet.merge_cells('Y1:Y2')
    sheet.merge_cells('Z1:Z2')
    sheet.merge_cells('AA1:AA2')
    cells_values_1 = [
        ('A1', '№'), ('B1', 'Дата и время'), ('C1', 'Адрес КЛ'), ('D1', '№ теста'),
        ('E1', 'Код 1'), ('F1', 'Код 2'), ('G1', 1), ('H1', 2), ('M1', 4),
        ('N1', 5), ('Q1', 7), ('T1', 8), ('U1', 9), ('X1', 'Резерв 1'),
        ('Y1', 'Резерв 2'), ('Z1', 'Описание 1'), ('AA1', 'Описание 2')
    ]
    for cell, value in cells_values_1:
        sheet[cell] = value
    cells_values_2 = [
        ('G2', 'Время готовности'), ('H2', 'Тип'), ('I2', "Версия"),
        ('J2', "Длит. СБ"), ('K2', "Ток СБ"), ('L2', "Средн. ток"),
        ('M2', "Т (С)"), ('N2', "Uв"), ('O2', "Uc"), ('P2', "Uн"),
        ('Q2', "Св"), ('R2', "Сс"), ('S2', "Сн"), ('T2', "Ток заряда"),
        ('U2', "Св"), ('V2', "Сс"), ('W2', "Сн")
    ]
    for cell, value in cells_values_2:
        sheet[cell] = value

    if not period:
        tests = get_unarchived_tests()
    else:
        if is_error_flag:
            tests = get_error_tests_for_period(period)
        else:
            tests = get_tests_for_period(period)

    for test in tests:
        str_code_1_2 = f'{test.exit_code_1}{test.exit_code_2}'
        if str_code_1_2 in ('73', '74', '92', '93'):
            str_code_3 = '0' * (3 - len(str(test.exit_code_3))) + str(test.exit_code_3)
        else:
            str_code_3 = test.exit_code_3
        params = [
            test.kl, test.exit_code_1, test.exit_code_2, str_code_3, test.ready_time,
            test.device_type, test.device_version, test.sb_duration, test.electricity_sb,
            test.average_electricity, test.temperature, test.Uv, test.Us, test.Un, test.Cv,
            test.Cs, test.Cn, test.charge_electricity_current, test.discharge_rate_upper_ionistor,
            test.discharge_rate_middle_ionistor, test.discharge_rate_lower_ionistor,
            test.reserve_1, test.reserve_2, test.description_1, test.description_2
        ]
        line = [test.number, test.time.strftime('%Y.%m.%d %H:%M:%S')]
        for param in params:
            line.append(param)
        sheet.append(line)

    for row in list(sheet.rows):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    for cell in ["C1", "G2", "J2", "K2", "L2", "L1", "T2"]:
        sheet[cell].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['G'].width = 12
    sheet.column_dimensions['Z'].width = 50
    sheet.column_dimensions['AA'].width = 25
    workbook.save(filename)
    workbook.close()


def write_report_to_file(filename, report, is_old=False):
    """Записывает отчет за весь день в эксель-файл.
    Если при прошлом сеансе отчет не был по какой-то
    причине автоматически сохранен, то данная функция будет
    запущена при новом сеансе, при этом будет передан
    флаг is_old = True."""
    date = report.date_time.strftime('%d-%m-%Y')
    try:
        new_workbook = Workbook()
        new_workbook.save(filename)
        workbook = load_workbook(filename)
        sheet = workbook.worksheets[0]
        sheet.merge_cells('A1:F1')
        sheet.merge_cells('A2:F2')
        sheet['A1'] = f'Отчет от {date}'
        sheet['A2'] = f'Всего тестов запущено: {report.success_all + report.error_all + report.interrupted_all}'
        sheet.append(['', 'ИПТ-И', 'ИПТ-СИ', 'МКС', "ИПД", "ИПР"])
        sheet.append(['Успешно', report.success_ipt, report.success_ipt_s, report.success_mks, report.success_ipd, report.success_ipr])
        sheet.append(['Ошибка', report.error_ipt, report.error_ipt_s, report.error_mks, report.error_ipd, report.error_ipr])

        for row in list(sheet.rows):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        sheet['A1'].alignment = Alignment(horizontal='left')
        sheet['A2'].alignment = Alignment(horizontal='left')
        workbook.save(filename)
        workbook.close()
    except Exception as error:
        raise Exception(error)


def write_report_for_period_to_file(filename, period):
    """Записывает отчет за выбранный период в эксель-файл."""
    reports = get_reports_for_period(period)
    success_all = sum([report.success_all for report in reports])
    error_all = sum([report.error_all for report in reports])
    interrupted_all = sum([report.interrupted_all for report in reports])
    all = success_all + error_all + interrupted_all
    success_ipt = sum([report.success_ipt for report in reports])
    success_ipt_s = sum([report.success_ipt_s for report in reports])
    success_mks = sum([report.success_mks for report in reports])
    success_ipd = sum([report.success_ipd for report in reports])
    success_ipr = sum([report.success_ipr for report in reports])
    error_ipt = sum([report.error_ipt for report in reports])
    error_ipt_s = sum([report.error_ipt_s for report in reports])
    error_mks = sum([report.error_mks for report in reports])
    error_ipd = sum([report.error_ipd for report in reports])
    error_ipr = sum([report.error_ipr for report in reports])

    new_date_to = period[1] - timedelta(days=1)  # для корректного отображения даты в таблице эксель, т.к. фильтрация идет не включая правое значение
    date_from = period[0].strftime('%d.%m.%Y')
    date_to = new_date_to.strftime('%d.%m.%Y')

    new_workbook = Workbook()
    new_workbook.save(filename)
    workbook = load_workbook(filename)
    sheet = workbook.worksheets[0]
    sheet.merge_cells('A1:F1')
    sheet.merge_cells('A2:F2')
    sheet['A1'] = f'Отчет за период c {date_from} по {date_to}'
    sheet['A2'] = f'Всего тестов запущено: {all}'
    sheet.append(['', 'ИПТ-И', 'ИПТ-СИ', 'МКС', "ИПД", "ИПР"])
    sheet.append(['Успешно', success_ipt, success_ipt_s, success_mks, success_ipd, success_ipr])
    sheet.append(['Ошибка', error_ipt, error_ipt_s, error_mks, error_ipd, error_ipr])

    for row in list(sheet.rows):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    sheet['A1'].alignment = Alignment(horizontal='left')
    sheet['A2'].alignment = Alignment(horizontal='left')
    workbook.save(filename)
    workbook.close()
